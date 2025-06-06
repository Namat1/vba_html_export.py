import streamlit as st
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
import zipfile
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

# Kalenderwoche berechnen
def get_calendar_week(date):
    return date.isocalendar()[1]

# Dateinamen säubern
def clean_filename(name):
    return "".join(c for c in name if c.isalnum() or c in "._- ").strip().replace(" ", "_")

# HTML für Fahrerzeile erzeugen (2 Touren pro Tag = 2 Zeilen mit Datum/Wochentag)
def create_html_from_row_full_repeat(ws, row_idx, base_date, kw, fahrer_name):
    tage = ["Sonntag", "Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag"]
    html = f"""<!DOCTYPE html>
<html lang='de'>
<head>
  <meta charset='UTF-8'>
  <meta name='viewport' content='width=device-width, initial-scale=1.0'>
  <title>KW{kw} – {fahrer_name}</title>
  <style>
    body{{font-family:Arial,sans-serif;margin:20px;}}
    h1,h2{{text-align:center;}}
    table{{width:100%;border-collapse:collapse;margin-top:20px;}}
    th,td{{border:1px solid #ddd;padding:8px;text-align:center;}}
    th{{background-color:#f4f4f4;}}
    @media(max-width:600px){{
      table,thead,tbody,th,td,tr{{display:block;}}
      th{{display:none;}}
      td{{border:none;border-bottom:1px solid #eee;position:relative;padding-left:50%;}}
      td:before{{position:absolute;top:8px;left:8px;width:45%;white-space:nowrap;font-weight:bold;}}
    }}
  </style>
</head>
<body>
  <h1>KW{kw}</h1>
  <h2>{fahrer_name} – Schichtbeginn: {base_date.strftime('%d.%m.%Y')} bis {(base_date + timedelta(days=6)).strftime('%d.%m.%Y')}</h2>
  <table><thead><tr><th>Datum</th><th>Wochentag</th><th>Uhrzeit</th><th>Tour / Aufgabe</th></tr></thead><tbody>"""

    for j in range(7):
        einsatz_count = 0
        tag_datum = base_date + timedelta(days=j)
        wochentag = tage[j]

        for k in range(2):  # max. 2 Einträge pro Tag
            uhrzeit_col = 5 + j * 2 + k
            tour = ws.cell(row=row_idx + 1, column=uhrzeit_col + 1).value
            uhrzeit = ws.cell(row=row_idx, column=uhrzeit_col + 1).value

            if str(tour).strip() not in ("", "0", "None"):
                if isinstance(uhrzeit, (int, float)):
                    if uhrzeit in [0, 1]:
                        uhrzeit_str = "00:00"
                    elif 0 < uhrzeit < 1:
                        t = from_excel(uhrzeit)
                        uhrzeit_str = t.strftime("%H:%M")
                    else:
                        uhrzeit_str = str(uhrzeit)
                elif isinstance(uhrzeit, datetime):
                    uhrzeit_str = uhrzeit.strftime("%H:%M")
                elif str(uhrzeit).strip() in ["", "0", "0.0", "1", "1.0", "00:00"]:
                    uhrzeit_str = "00:00"
                else:
                    uhrzeit_str = str(uhrzeit)

                html += f"<tr><td>{tag_datum.strftime('%d.%m.%Y')}</td><td>{wochentag}</td><td>{uhrzeit_str}</td><td>{tour}</td></tr>\n"
                einsatz_count += 1

        if einsatz_count == 0:
            html += f"<tr><td>{tag_datum.strftime('%d.%m.%Y')}</td><td>{wochentag}</td><td></td><td></td></tr>\n"

    html += "</tbody></table></body></html>"
    return html

# Hauptverarbeitung: HTML-Dateien erzeugen und zippen
def process_excel_full_repeat(excel_bytes):
    in_memory_zip = BytesIO()
    with tempfile.TemporaryDirectory() as tmpdirname:
        wb = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
        ws = wb["Druck Fahrer"]

        datum = ws["E2"].value
        if not isinstance(datum, datetime):
            datum = pd.to_datetime(datum)

        kw = get_calendar_week(datum)
        name_dict = {}

        i = 11
        while i <= ws.max_row:
            fahrer_name = ws[f"B{i}"].value
            if fahrer_name:
                base_name = fahrer_name.strip()
                count = name_dict.get(base_name, 0)
                name_dict[base_name] = count + 1
                final_name = base_name if count == 0 else f"{base_name}_{count}"
                safe_name = clean_filename(final_name)

                html = create_html_from_row_full_repeat(ws, i, datum, kw, final_name)
                html_path = Path(tmpdirname) / f"KW{kw}_{safe_name}.html"
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(html)

                i += 2
            else:
                i += 1

        with zipfile.ZipFile(in_memory_zip, "w", zipfile.ZIP_DEFLATED) as zipf:
            for html_file in Path(tmpdirname).glob("*.html"):
                zipf.write(html_file, html_file.name)

    in_memory_zip.seek(0)
    return in_memory_zip

# Streamlit App-UI
st.title("Fahrer-Wochenplaner (Excel zu HTML)")

uploaded_file = st.file_uploader("Excel-Datei hochladen (Blatt: 'Druck Fahrer')", type=["xlsx", "xlsm"])

if uploaded_file:
    try:
        zip_bytes = process_excel_full_repeat(uploaded_file.read())
        st.success("Export erfolgreich!")
        st.download_button(
            label="ZIP-Datei herunterladen",
            data=zip_bytes,
            file_name="html_export.zip",
            mime="application/zip"
        )
    except Exception as e:
        st.error(f"Fehler beim Verarbeiten der Datei: {e}")
